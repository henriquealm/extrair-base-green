import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_COLS = {
    "email":             "E-mail",
    "fullName":          "Nome Completo",
    "segment":           "Segmento",
    "all_categories":    "Categorias",
    "isNewsletterOptIn": "Newsletter",
    "_source":           "Origem do Dado",
}

SEGMENT_LABELS = {
    "meninos":    "🔵 Somente Meninos",
    "meninas":    "🩷 Somente Meninas",
    "ambos":      "💜 Meninos e Meninas",
    "neutro":     "⚪ Categorias Neutras",
    "sem_compra": "❌ Sem Compra",
}

SEGMENT_COLORS = {
    "meninos":    ("1F4E79", "DBEAFE"),
    "meninas":    ("C9184A", "FCE4EC"),
    "ambos":      ("6A1B9A", "EDE7F6"),
    "neutro":     ("424242", "F5F5F5"),
    "sem_compra": ("B71C1C", "FFEBEE"),
}

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _style_header(ws, n_cols: int, color: str):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=1, column=col)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = BORDER
    ws.row_dimensions[1].height = 28


def _write_df_to_sheet(ws, df: pd.DataFrame, header_color: str, row_colors: tuple):
    headers = list(OUTPUT_COLS.values())
    cols    = list(OUTPUT_COLS.keys())

    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h)

    _style_header(ws, len(headers), header_color)

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        bg = row_colors[(row_idx) % 2]
        for col_idx, col in enumerate(cols, 1):
            val = row.get(col, "")
            if col == "isNewsletterOptIn":
                val = "Sim" if str(val).lower() == "true" else "Não"
            if col == "segment":
                val = SEGMENT_LABELS.get(str(val), val)
            c = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
            c.font      = Font(name="Arial", size=9)
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(vertical="center")
            c.border    = BORDER
        ws.row_dimensions[row_idx].height = 17

    widths = [36, 20, 24, 52, 12, 16]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def export_to_excel(df: pd.DataFrame, path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)

    writer = pd.ExcelWriter(path, engine="openpyxl")
    # cria o arquivo com uma sheet vazia para depois manipular
    pd.DataFrame().to_excel(writer, sheet_name="temp", index=False)
    writer.close()

    wb = load_workbook(path)
    if "temp" in wb.sheetnames:
        del wb["temp"]

    # ── Aba Resumo ──────────────────────────────
    ws_sum = wb.create_sheet("📊 Resumo", 0)
    _build_summary(ws_sum, df)

    # ── Abas por segmento ───────────────────────
    for seg, label in SEGMENT_LABELS.items():
        hdr_color, row1, row2 = SEGMENT_COLORS[seg][0], SEGMENT_COLORS[seg][1], "FFFFFF"
        seg_df = df[df["segment"] == seg].copy()
        ws     = wb.create_sheet(label)
        _write_df_to_sheet(ws, seg_df, hdr_color, (row1, row2))

    # ── Aba Todos com compra ─────────────────────
    buyers_df = df[df["segment"] != "sem_compra"].copy()
    ws_all    = wb.create_sheet("📦 Todos com Compra")
    _write_df_to_sheet(ws_all, buyers_df, "1B5E20", ("E8F5E9", "FFFFFF"))

    wb.save(path)
    print(f"💾 Salvo: {path}")


def _build_summary(ws, df: pd.DataFrame):
    # título
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "SEGMENTAÇÃO DE CLIENTES — GREEN BY MISSAKO"
    c.font      = Font(bold=True, size=14, color="FFFFFF", name="Arial")
    c.fill      = PatternFill("solid", start_color="1A1A2E")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    headers = ["Segmento", "Total", "% da Base Total", "% dos Compradores", "Origem Principal"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color="2D2D44")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = BORDER
    ws.row_dimensions[2].height = 24

    total       = len(df)
    total_buyers = len(df[df["segment"] != "sem_compra"])

    rows = [
        ("meninos",    "MD Tag + OMS"),
        ("meninas",    "MD Tag + OMS"),
        ("ambos",      "MD Tag + OMS"),
        ("neutro",     "MD Tag + OMS"),
        ("sem_compra", "Apenas cadastro"),
    ]

    for row_idx, (seg, origem) in enumerate(rows, 3):
        label  = SEGMENT_LABELS[seg]
        count  = len(df[df["segment"] == seg])
        pct_b  = f"{count/total*100:.1f}%" if total else "—"
        pct_c  = f"{count/total_buyers*100:.1f}%" if total_buyers and seg != "sem_compra" else "—"
        bg     = SEGMENT_COLORS[seg][1]

        vals = [label, count, pct_b, pct_c, origem]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font      = Font(name="Arial", size=10, bold=(col_idx == 1))
            c.fill      = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(
                horizontal="center" if col_idx > 1 else "left",
                vertical="center"
            )
            c.border = BORDER
        ws.row_dimensions[row_idx].height = 22

    # totais
    row_idx = len(rows) + 3
    for col_idx, val in enumerate(["TOTAL GERAL", total, "100%", f"{total_buyers} compradores", ""], 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font      = Font(bold=True, name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color="E8F5E9")
        c.alignment = Alignment(horizontal="center" if col_idx > 1 else "left", vertical="center")
        c.border    = BORDER
    ws.row_dimensions[row_idx].height = 22

    widths = [30, 10, 18, 20, 20]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w